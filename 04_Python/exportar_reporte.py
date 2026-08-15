import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Crear una base de datos limpia e independiente en la memoria RAM
conn = sqlite3.connect(':memory:')
cursor = conn.cursor()

# 2. Cargar tablas y datos desde tu archivo .sql
with open('01_SQL/01_asadero_el_titular.sql', 'r', encoding='utf-8') as file:
    sql_script = file.read()

cursor.executescript(sql_script)

# 3. Consulta SQL de Ingeniería de Menú
query = """
SELECT 
    p.nombre_producto,
    COALESCE(SUM(v.cantidad), 0) AS unidades_vendidas,
    COALESCE(SUM(v.cantidad * p.precio_venta), 0) AS ingreso_total,
    COALESCE(SUM(v.cantidad * (p.precio_venta - p.costo_estimado)), 0) AS ganancia_bruta,
    CASE 
        WHEN SUM(v.cantidad * p.precio_venta) > 0 
        THEN SUM(v.cantidad * (p.precio_venta - p.costo_estimado)) * 1.0 / SUM(v.cantidad * p.precio_venta)
        ELSE 0 
    END AS margen_porcentaje
FROM productos p
LEFT JOIN ventas_detalle v ON p.id_producto = v.id_producto
GROUP BY p.id_producto, p.nombre_producto
ORDER BY ganancia_bruta DESC;
"""

df = pd.read_sql_query(query, conn)

# 4. Generar el Excel base
excel_path = '04_Python/reporte_ingenieria_menu.xlsx'
df.to_excel(excel_path, index=False)

# =========================================================
# 5. DISEÑO Y FORMATO PROFESIONAL (OPENPYXL)
# =========================================================
wb = openpyxl.load_workbook(excel_path)
ws = wb.active
ws.title = "Ingeniería de Menú"

# Estilos visuales
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Azul ejecutivo
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

encabezados = ["Producto", "Unidades Vendidas", "Ingreso Total", "Ganancia Bruta", "Margen %"]

# Dar formato a los Encabezados (Fila 1)
for col_idx, text in enumerate(encabezados, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = text
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Dar formato a los datos (Filas 2 en adelante)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
    row[0].alignment = Alignment(horizontal="left")      # Producto
    row[1].number_format = '#,##0'                       # Unidades
    row[1].alignment = Alignment(horizontal="center")
    row[2].number_format = '$#,##0.00'                   # Ingreso Total ($)
    row[3].number_format = '$#,##0.00'                   # Ganancia Bruta ($)
    row[4].number_format = '0.0%'                        # Margen (%)
    
    for cell in row:
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=10)

# Autoajustar el ancho de las columnas
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

# Guardar el archivo formateado
wb.save(excel_path)

print("✨ ¡Reporte generado y formateado con éxito sin errores!")
